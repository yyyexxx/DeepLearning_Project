"""Excel 报销单导出。"""

from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


EXPORT_FIELDS = [
    ("发票代码", "invoice_code"),
    ("发票号码", "invoice_number"),
    ("开票日期", "invoice_date"),
    ("购买方", "purchaser"),
    ("销售方", "seller"),
    ("不含税金额", "amount"),
    ("税额", "tax"),
    ("价税合计", "total"),
    ("报销人", "reimburser"),
    ("报销事由", "reason"),
    ("所属部门", "department"),
    ("提交日期", "submitted_date"),
]


def generate_excel(data: dict) -> BytesIO:
    """根据报销数据生成 Excel 文件，返回 BytesIO。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报销单"

    # 标题行
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "费用报销单"
    title_cell.font = Font(name="微软雅黑", size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 字段
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    row = 3
    for label, key in EXPORT_FIELDS:
        # 标签列
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name="微软雅黑", size=11, bold=True)
        label_cell.fill = header_fill
        label_cell.border = thin_border
        label_cell.alignment = Alignment(horizontal="right", vertical="center")

        # 值列（合并 B-D）
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        value_cell = ws.cell(row=row, column=2, value=_fmt(data, key))
        value_cell.font = Font(name="微软雅黑", size=11)
        value_cell.border = thin_border
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row].height = 25
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _fmt(data: dict, key: str) -> str:
    val = data.get(key)
    if val is None:
        return ""
    if key in ("amount", "tax", "total") and isinstance(val, (int, float)):
        return f"¥{val:,.2f}"
    return str(val)
