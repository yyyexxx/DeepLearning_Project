"""测试导出模块：Excel 和 PDF 生成。"""

import pytest
from openpyxl import load_workbook

from export.excel_writer import generate_excel
from export.pdf_writer import generate_pdf


TEST_DATA = {
    "invoice_code": "123456789012",
    "invoice_number": "87654321",
    "invoice_date": "2024-01-15",
    "purchaser": "北京测试科技有限公司",
    "seller": "上海数据服务有限公司",
    "amount": 1234.56,
    "tax": 160.49,
    "total": 1395.05,
    "reimburser": "张三",
    "reason": "差旅报销",
    "department": "技术部",
    "submitted_date": "2024-02-01",
}


class TestExcelExport:
    def test_generates_valid_excel(self):
        buf = generate_excel(TEST_DATA)
        assert buf is not None
        assert len(buf.getvalue()) > 500

    def test_excel_contains_data(self):
        buf = generate_excel(TEST_DATA)
        wb = load_workbook(buf)
        ws = wb.active
        # 标题
        assert ws["A1"].value == "费用报销单"
        # 找个字段验证
        all_text = ""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    all_text += str(cell.value)
        assert "123456789012" in all_text
        assert "张三" in all_text


class TestPDFExport:
    def test_generates_valid_pdf(self):
        buf = generate_pdf(TEST_DATA)
        pdf_bytes = buf.getvalue()
        assert pdf_bytes[:5] == b"%PDF-"

    def test_pdf_non_empty(self):
        buf = generate_pdf(TEST_DATA)
        assert len(buf.getvalue()) > 500
